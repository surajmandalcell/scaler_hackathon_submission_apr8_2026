"""Excel template generation and bulk-upload parsing for FundLens."""
from __future__ import annotations

import tempfile
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Styling helpers ───────────────────────────────────────────────────────────

_HEADER_FILL  = PatternFill("solid", fgColor="0F3460")
_EXAMPLE_FILL = PatternFill("solid", fgColor="1E293B")
_HEADER_FONT  = Font(bold=True, color="E2E8F0", size=11)
_EXAMPLE_FONT = Font(color="94A3B8", size=10, italic=True)
_BODY_FONT    = Font(color="E2E8F0", size=10)
_THIN_BORDER  = Border(bottom=Side(style="thin", color="334155"))

def _style_header(ws, headers: list[str], col_widths: list[int]) -> None:
    ws.append(headers)
    for i, (cell, w) in enumerate(zip(ws[1], col_widths), start=1):
        cell.font       = _HEADER_FONT
        cell.fill       = _HEADER_FILL
        cell.alignment  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 32

def _style_row(ws, row_idx: int, is_example: bool = False) -> None:
    for cell in ws[row_idx]:
        cell.font      = _EXAMPLE_FONT if is_example else _BODY_FONT
        cell.fill      = _EXAMPLE_FILL
        cell.border    = _THIN_BORDER
        cell.alignment = Alignment(vertical="center")

def _note_cell(ws, cell_ref: str, note: str) -> None:
    cell = ws[cell_ref]
    cell.value     = note
    cell.font      = Font(color="64748B", size=9, italic=True)
    cell.fill      = PatternFill("solid", fgColor="0F172A")
    cell.alignment = Alignment(horizontal="left", vertical="center")


# ── Onboarding template ───────────────────────────────────────────────────────

def generate_onboarding_template() -> str:
    """Two-sheet workbook: Funds + Investments."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Funds ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Funds"
    ws1.sheet_view.showGridLines = False

    _style_header(ws1,
        ["Fund Name", "Onboarded Quarter / Date"],
        [36, 28])

    _note_cell(ws1, "A2", "↓ Example rows (delete before uploading)")

    for i, row in enumerate([
        ("RE Alpha Fund I",   "Q1 2022"),
        ("RE Beta Fund II",   "Q3 2023"),
        ("RE Gamma Fund III", "Q2 2024"),
    ], start=3):
        ws1.append(row)
        _style_row(ws1, i, is_example=True)

    # ── Sheet 2: Investments ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Investments")
    ws2.sheet_view.showGridLines = False

    _style_header(ws2,
        ["Deal Name", "Fund Name (investing fund)",
         "Fund's Agreed Ownership %", "Onboarded Quarter / Date"],
        [32, 30, 26, 28])

    _note_cell(ws2, "A2",
               "↓ Example rows — one row per deal-fund relationship. "
               "Co-investment = same deal, different fund rows.")

    inv_rows: list[tuple[Any, ...]] = [
        ("Embassy Office Park",  "RE Alpha Fund I",   100, "Q1 2022"),
        ("Prestige Residences",  "RE Alpha Fund I",   100, "Q2 2022"),
        ("Prestige Tower",       "RE Beta Fund II",    40, "Q3 2023"),
        ("Prestige Tower",       "RE Gamma Fund III",  35, "Q3 2023"),
    ]
    for i, inv_row in enumerate(inv_rows, start=3):
        ws2.append(inv_row)
        _style_row(ws2, i, is_example=True)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx",
                                     prefix="fundlens_onboarding_") as f:
        wb.save(f.name)
        return f.name


# ── Cashflow template ─────────────────────────────────────────────────────────

def generate_cashflow_template() -> str:
    """
    Simplified cashflow upload template.
    Columns: Fund Name | Deal Name | Date | Transaction Type | Deal Amount | Fund Amount (USD M)

    All reporting at fund level in USD — no FX column, no ROC, no split disposition.
    Disposition is a single combined transaction (ROC + G/L on investment + G/L on FX all rolled in).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cashflows"
    ws.sheet_view.showGridLines = False

    _style_header(ws, [
        "Fund Name",
        "Deal Name",
        "Date (YYYY-MM-DD)",
        "Transaction Type",
        "Deal Amount (deal currency)\nOptional — for your records",
        "Fund Amount (USD M)\nRequired — stored in DB",
    ], [28, 28, 22, 26, 32, 28])

    ws.row_dimensions[1].height = 40

    _note_cell(ws, "A2",
               "Transaction Types:  Investment  |  Current Income  |  Disposition  —  "
               "Investment = capital call (Fund Amount NEGATIVE, e.g. -1.5).  "
               "Current Income = rent/yield (positive).  "
               "Disposition = sale proceeds all-in (positive — ROC + G/L combined, no split needed).")

    for i, row in enumerate([
        ("RE Alpha Fund I", "Embassy Office Park",  "2022-01-15", "Investment",    -8_000_000, -8.0),
        ("RE Alpha Fund I", "Embassy Office Park",  "2023-06-01", "Investment",    -3_000_000, -3.0),
        ("RE Alpha Fund I", "Embassy Office Park",  "2024-09-30", "Disposition",    2_500_000,  2.5),
        ("RE Alpha Fund I", "Embassy Office Park",  "2024-12-31", "Current Income",  400_000,   0.4),
        ("RE Beta Fund II", "Prestige Tower",       "2021-09-01", "Investment",    -4_000_000, -4.0),
        ("RE Beta Fund II", "Prestige Tower",       "2024-08-01", "Disposition",    1_800_000,  1.8),
        ("RE Beta Fund II", "Prestige Tower",       "2024-12-31", "Current Income",  320_000,   0.32),
    ], start=3):
        ws.append(row)
        _style_row(ws, i, is_example=True)
        # Format the Fund Amount column as number
        ws.cell(i, 6).number_format = '#,##0.0000"M"'

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx",
                                     prefix="fundlens_cashflows_") as f:
        wb.save(f.name)
        return f.name


# ── Quarterly NAV template ────────────────────────────────────────────────────

def generate_quarterly_template() -> str:
    """Single-sheet quarterly NAV template (Beginning NAV per fund per quarter)."""
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Quarterly NAV"
    ws.sheet_view.showGridLines = False

    _style_header(ws,
        ["Fund Name", "Quarter End Date", "Quarter Start Date",
         "Beginning NAV (USD M)"],
        [30, 22, 22, 26])

    _note_cell(ws, "A2",
               "Beginning NAV = prior quarter's closing appraiser value.  "
               "Ending NAV is computed automatically from deal-level appraiser values — do not enter here.")

    for i, row in enumerate([
        ("RE Alpha Fund I", "2024-03-31", "2024-01-01", 280.0),
        ("RE Alpha Fund I", "2024-06-30", "2024-04-01", 295.5),
        ("RE Alpha Fund I", "2024-09-30", "2024-07-01", 302.0),
        ("RE Alpha Fund I", "2024-12-31", "2024-10-01", 308.2),
    ], start=3):
        ws.append(row)
        _style_row(ws, i, is_example=True)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx",
                                     prefix="fundlens_quarterly_") as f:
        wb.save(f.name)
        return f.name


# ── Upload parsers ────────────────────────────────────────────────────────────

def _val(v: Any) -> Any:
    return v.strip() if isinstance(v, str) else v


def parse_onboarding_upload(filepath: str) -> tuple[list, list, str | None]:
    """Parse onboarding template. Returns (fund_dicts, inv_dicts, error_str)."""
    try:
        wb    = openpyxl.load_workbook(filepath, data_only=True)
        funds, investments = [], []

        if "Funds" in wb.sheetnames:
            for row in wb["Funds"].iter_rows(min_row=2, values_only=True):
                name = _val(row[0]) if len(row) > 0 else None
                qtr  = _val(row[1]) if len(row) > 1 else None
                if name and "↓" not in (name or ""):
                    funds.append({"name": name, "quarter": qtr or ""})

        if "Investments" in wb.sheetnames:
            for row in wb["Investments"].iter_rows(min_row=2, values_only=True):
                dname = _val(row[0]) if len(row) > 0 else None
                fname = _val(row[1]) if len(row) > 1 else None
                opct  = row[2]       if len(row) > 2 else None
                qtr   = _val(row[3]) if len(row) > 3 else None
                if dname and fname and "↓" not in (dname or ""):
                    investments.append({
                        "deal_name":     dname,
                        "fund_name":     fname,
                        "ownership_pct": float(opct) if opct is not None else 100.0,
                        "quarter":       qtr or "",
                    })

        return funds, investments, None
    except Exception as exc:
        return [], [], str(exc)


def parse_cashflow_upload(filepath: str) -> tuple[list, str | None]:
    """
    Parse cashflow template.
    Columns: Fund Name | Deal Name | Date | Transaction Type | Deal Amount (optional) | Fund Amount (USD M)
    Returns (cf_dicts, error_str).
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        cfs = []

        if "Cashflows" not in wb.sheetnames:
            return [], "Sheet 'Cashflows' not found"

        for row in wb["Cashflows"].iter_rows(min_row=2, values_only=True):
            fname    = _val(row[0]) if len(row) > 0 else None
            dname    = _val(row[1]) if len(row) > 1 else None
            date     = str(_val(row[2])).split(" ")[0] if len(row) > 2 and row[2] else None
            txn_type = _val(row[3]) if len(row) > 3 else None
            # col 4: deal amount (optional, ignored for storage)
            fund_amt = row[5]       if len(row) > 5 else None

            if not fname or "↓" in (fname or ""):
                continue
            if fname and dname and date and txn_type and fund_amt is not None:
                cfs.append({
                    "fund_name": fname,
                    "deal_name": dname,
                    "date":      date,
                    "txn_type":  txn_type,
                    "fund_amt":  float(fund_amt),
                })

        return cfs, None
    except Exception as exc:
        return [], str(exc)


def parse_quarterly_upload(filepath: str) -> tuple[list, list, str | None]:
    """Parse quarterly NAV template. Returns (nav_dicts, [], error_str)."""
    try:
        wb   = openpyxl.load_workbook(filepath, data_only=True)
        navs = []

        if "Quarterly NAV" in wb.sheetnames:
            for row in wb["Quarterly NAV"].iter_rows(min_row=2, values_only=True):
                fname  = _val(row[0]) if len(row) > 0 else None
                qend   = str(_val(row[1])).split(" ")[0] if len(row) > 1 and row[1] else None
                qstart = str(_val(row[2])).split(" ")[0] if len(row) > 2 and row[2] else None
                bnav   = row[3] if len(row) > 3 else None
                if fname and qend and "↓" not in (fname or ""):
                    navs.append({
                        "fund_name":    fname,
                        "quarter_end":  qend,
                        "quarter_start": qstart or "",
                        "begin_nav":    float(bnav) if bnav is not None else None,
                    })

        return navs, [], None
    except Exception as exc:
        return [], [], str(exc)
