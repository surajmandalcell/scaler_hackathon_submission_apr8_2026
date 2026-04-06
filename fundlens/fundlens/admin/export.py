"""Excel export — computed answer key for admin reconciliation."""
from __future__ import annotations
import tempfile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fundlens.server.calculations import compute_nav_bridge, compute_metrics


BLUE   = "0F3460"
LBLUE  = "1E3A5F"
GREEN  = "0F4C35"
LBGBLU = "D6E4F0"
YELLOW = "FFF2CC"
ALT    = "F0F4F8"

thin = Side(style="thin", color="BBBBBB")
brd  = Border(left=thin, right=thin, top=thin, bottom=thin)


def _hdr(ws, r, c, text, bg=LBLUE, fg="FFFFFF"):
    cell = ws.cell(r, c, text)
    cell.font = Font(bold=True, color=fg, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = brd
    return cell


def _val(ws, r, c, v, fmt=None, bold=False, bg=None, align="left"):
    cell = ws.cell(r, c, v)
    cell.font = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = brd
    if fmt:
        cell.number_format = fmt
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    return cell


def export_answer_key(store) -> str:
    """
    Generate an Excel workbook with:
      Sheet 1 — Summary across all funds
      Sheet per fund — raw cashflows + 8-line NAV bridge + metrics
    Returns path to the temp .xlsx file.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Summary ──────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    ws_sum.column_dimensions["A"].width = 3
    for col, w in zip("BCDEFGH", [28, 16, 16, 12, 12, 12, 14]):
        ws_sum.column_dimensions[col].width = w

    ws_sum.cell(1, 2, "FundLens — Computed Answer Key").font = Font(bold=True, size=14, color=BLUE)
    ws_sum.cell(2, 2, "Generated from DB data. Use this to reconcile your Excel.").font = \
        Font(italic=True, size=10, color="666666")

    row = 4
    _hdr(ws_sum, row, 2, "Fund Name")
    _hdr(ws_sum, row, 3, "Beginning NAV (USD M)")
    _hdr(ws_sum, row, 4, "Ending NAV (USD M)")
    _hdr(ws_sum, row, 5, "MOIC")
    _hdr(ws_sum, row, 6, "IRR")
    _hdr(ws_sum, row, 7, "Write Up/Down (Plug)")
    ws_sum.row_dimensions[row].height = 30
    row += 1

    for i, (fid, fund) in enumerate(store.funds.items()):
        bridge  = compute_nav_bridge(fid, store)
        metrics = compute_metrics(fid, store)
        bg = ALT if i % 2 else None
        _val(ws_sum, row, 2, fund.fund_name, bold=True, bg=bg)
        _val(ws_sum, row, 3, fund.beginning_nav,              "#,##0.0000", bg=bg, align="right")
        _val(ws_sum, row, 4, fund.ending_nav,                 "#,##0.0000", bg=bg, align="right")
        _val(ws_sum, row, 5, metrics.get("moic", ""),         '0.0000"x"',  bg=bg, align="right")
        _val(ws_sum, row, 6, metrics.get("irr", ""),          "0.00%",       bg=bg, align="right")
        _val(ws_sum, row, 7, bridge.get("write_up_down", ""), "#,##0.0000", bg=bg, align="right")
        row += 1

    # ── One sheet per fund ────────────────────────────────────────────────
    bridge_labels = [
        ("beginning_nav",         "Beginning NAV",              False, LBGBLU),
        ("contribution",          "(+) Contribution",           False, None),
        ("disposition",           "(−) Disposition",            False, None),
        ("income",                "(+) Income",                 False, None),
        ("cashflow_adjusted_nav", "= CF Adjusted NAV",          True,  LBGBLU),
        ("income_reversal",       "(−) Income Reversal",        False, None),
        ("write_up_down",         "(+/−) Write Up/Down (PLUG)", False, YELLOW),
        ("ending_nav",            "= Ending NAV",               True,  LBGBLU),
    ]

    for fid, fund in store.funds.items():
        name = fund.fund_name[:31]  # Excel sheet name max 31 chars
        ws = wb.create_sheet(name)
        for col, w in zip("ABCDEFG", [3, 32, 16, 16, 14, 14, 14]):
            ws.column_dimensions[col].width = w

        # Title
        ws.cell(1, 2, f"{fund.fund_name} — Answer Key").font = \
            Font(bold=True, size=13, color=BLUE)
        ws.cell(2, 2, f"Reporting: {fund.reporting_date}  |  "
                      f"Period: {fund.nav_period_start or '—'} → {fund.reporting_date}").font = \
            Font(italic=True, size=10, color="555555")

        row = 4

        # ── Raw Cashflows ─────────────────────────────────────────────────
        ws.cell(row, 2, "Raw Cashflows (source data)").font = \
            Font(bold=True, size=11, color=BLUE)
        ws.cell(row, 2).fill = PatternFill("solid", fgColor=LBGBLU)
        row += 1

        _hdr(ws, row, 2, "Date")
        _hdr(ws, row, 3, "Deal")
        _hdr(ws, row, 4, "Type")
        _hdr(ws, row, 5, "Fund Amt (USD M)")
        cf_start = row + 1
        row += 1

        cfs = sorted(store.get_cashflows(fund_id=fid), key=lambda c: c.cash_date)
        for i, c in enumerate(cfs):
            deal = store.deals.get(c.deal_id)
            bg = ALT if i % 2 else None
            _val(ws, row, 2, c.cash_date,                          bg=bg, align="center")
            _val(ws, row, 3, deal.property_name if deal else c.deal_id, bg=bg)
            _val(ws, row, 4, c.cf_type,                            bg=bg)
            _val(ws, row, 5, c.fund_amt, "#,##0.0000",             bg=bg, align="right")
            row += 1

        row += 1

        # ── NAV Bridge ────────────────────────────────────────────────────
        ws.cell(row, 2, "8-Line NAV Bridge (computed)").font = \
            Font(bold=True, size=11, color=BLUE)
        ws.cell(row, 2).fill = PatternFill("solid", fgColor=LBGBLU)
        row += 1

        bridge = compute_nav_bridge(fid, store)
        _hdr(ws, row, 2, "Line Item")
        _hdr(ws, row, 3, "Computed Value (USD M)")
        _hdr(ws, row, 4, "Notes")
        row += 1

        notes = {
            "beginning_nav":         "Given (appraiser, prior quarter)",
            "contribution":          "= −Σ contributions in period",
            "disposition":           "= Σ dispositions in period",
            "income":                "= Σ income in period",
            "cashflow_adjusted_nav": "= Beg + Contribution − Disposition + Income",
            "income_reversal":       "= −Income (add back, income ≠ valuation change)",
            "write_up_down":         "= Ending − (CF Adj + Income Reversal)  ← PLUG",
            "ending_nav":            "Given (appraiser, current quarter)",
        }

        for key, label, bold, bg in bridge_labels:
            v = bridge.get(key, "")
            _val(ws, row, 2, label,  bold=bold, bg=bg or "FFFFFF")
            _val(ws, row, 3, v,      "#,##0.0000", bold=bold, bg=bg or "FFFFFF", align="right")
            _val(ws, row, 4, notes.get(key, ""), bg=bg or "FFFFFF")
            row += 1

        row += 1

        # ── Performance Metrics ───────────────────────────────────────────
        ws.cell(row, 2, "Performance Metrics (computed)").font = \
            Font(bold=True, size=11, color=BLUE)
        ws.cell(row, 2).fill = PatternFill("solid", fgColor=LBGBLU)
        row += 1

        metrics = compute_metrics(fid, store)
        _hdr(ws, row, 2, "Metric")
        _hdr(ws, row, 3, "Value")
        _hdr(ws, row, 4, "Formula")
        row += 1

        cfs_all = store.get_cashflows(fund_id=fid)
        total_contrib = sum(abs(c.fund_amt) for c in cfs_all if c.cf_type == "contribution")
        total_disp    = sum(c.fund_amt for c in cfs_all if c.cf_type == "disposition")
        total_income  = sum(c.fund_amt for c in cfs_all if c.cf_type == "income")

        _val(ws, row, 2, "Total Contributions (ITD)", bold=True)
        _val(ws, row, 3, total_contrib, "#,##0.0000", align="right")
        _val(ws, row, 4, "Σ |fund_amt| where cf_type = contribution")
        row += 1

        _val(ws, row, 2, "Total Dispositions (ITD)", bold=True)
        _val(ws, row, 3, total_disp, "#,##0.0000", align="right")
        _val(ws, row, 4, "Σ fund_amt where cf_type = disposition")
        row += 1

        _val(ws, row, 2, "Total Income (ITD)", bold=True)
        _val(ws, row, 3, total_income, "#,##0.0000", align="right")
        _val(ws, row, 4, "Σ fund_amt where cf_type = income")
        row += 1

        _val(ws, row, 2, "Ending NAV (Unrealized)", bold=True)
        _val(ws, row, 3, fund.ending_nav, "#,##0.0000", align="right")
        _val(ws, row, 4, "Appraiser value (given)")
        row += 1

        _val(ws, row, 2, "MOIC", bold=True, bg=LBGBLU)
        _val(ws, row, 3, metrics.get("moic", ""), '0.0000"x"', bold=True, bg=LBGBLU, align="right")
        _val(ws, row, 4, "(Dispositions + Income + Ending NAV) / Contributions", bg=LBGBLU)
        row += 1

        _val(ws, row, 2, "IRR", bold=True, bg=LBGBLU)
        _val(ws, row, 3, metrics.get("irr", ""), "0.00%", bold=True, bg=LBGBLU, align="right")
        _val(ws, row, 4, "XIRR on all dated cashflows + terminal Ending NAV", bg=LBGBLU)

    path = tempfile.mktemp(suffix=".xlsx", prefix="fundlens_answer_key_")
    wb.save(path)
    return path
