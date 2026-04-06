"""Generate FundLens Sample Data.xlsx — all 3 tasks, for validation & reconciliation."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fundlens.server.seed_data import load_easy_task, load_medium_task, load_hard_task, get_correct_answers
from fundlens.server.data_store import DataStore

store = DataStore()

# ── Style helpers ──────────────────────────────────────────────────────────
BLUE   = "1F4E79"
LBLUE  = "2E75B6"
LBGBLU = "D6E4F0"
LBGGRN = "E2EFDA"
YELLOW = "FFF2CC"
ALT    = "EBF5FB"

def hdr(ws, r, c, text, bg=LBLUE, fg="FFFFFF", wrap=False):
    cell = ws.cell(r, c, text)
    cell.font = Font(bold=True, color=fg, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    return cell

def val(ws, r, c, v, fmt=None, bold=False, bg=None, align="left"):
    cell = ws.cell(r, c, v)
    cell.font = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fmt: cell.number_format = fmt
    if bg:  cell.fill = PatternFill("solid", fgColor=bg)
    return cell

def sec(ws, r, c, text):
    cell = ws.cell(r, c, text)
    cell.font = Font(bold=True, size=11, color=BLUE)
    cell.fill = PatternFill("solid", fgColor=LBGBLU)
    return cell

thin = Side(style="thin", color="BBBBBB")
brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

def borders(ws, r1, r2, c1, c2):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            ws.cell(r, c).border = brd

def col_w(ws, mapping):
    for col, w in mapping.items():
        ws.column_dimensions[col].width = w


# ══════════════════════════════════════════════════════════════════════════
def build_task_sheet(wb, task_id, loader):
    store.clear()
    loader(store)
    answers = get_correct_answers(store)

    label = {"easy": "Level 1 — Easy", "medium": "Level 2 — Medium", "hard": "Level 3 — Hard"}[task_id]
    ws = wb.create_sheet(label)
    col_w(ws, {"A":3,"B":28,"C":14,"D":14,"E":14,"F":14,"G":16,"H":14,"I":16,"J":14})

    row = 1
    # ── Title ──────────────────────────────────────────────────────────────
    ws.cell(row, 2, f"FundLens Sample Data  ·  {label}").font = Font(bold=True, size=13, color=BLUE)
    row += 2

    # ── Fund summary ───────────────────────────────────────────────────────
    sec(ws, row, 2, "Funds")
    row += 1
    for h, c in [("Fund ID","B"),("Fund Name","C"),("Beg NAV (USD M)","D"),("End NAV (USD M)","E"),("Reporting Date","F")]:
        hdr(ws, row, "BCDEF".index(c)+2, h)
    row += 1
    for fid, fund in store.funds.items():
        val(ws, row, 2, fid)
        val(ws, row, 3, fund.fund_name)
        val(ws, row, 4, fund.beginning_nav, "#,##0.00", align="right")
        val(ws, row, 5, fund.ending_nav,    "#,##0.00", align="right")
        val(ws, row, 6, fund.reporting_date, align="center")
        row += 1
    row += 1

    # ── Cashflows ──────────────────────────────────────────────────────────
    sec(ws, row, 2, "Cashflows  (contribution = negative outflow | disposition = positive inflow | income = positive)")
    row += 1
    for c, h in enumerate(["Deal ID","Fund ID","Date","Type","Fund Amt (USD M)"], 2):
        hdr(ws, row, c, h, wrap=True)
    ws.row_dimensions[row].height = 28
    cf_start = row + 1
    for i, cf in enumerate(sorted(store.cashflows, key=lambda x: (x.fund_id, x.cash_date))):
        bg = ALT if i % 2 else None
        row += 1
        val(ws, row, 2, cf.deal_id,   bg=bg)
        val(ws, row, 3, cf.fund_id,   bg=bg)
        val(ws, row, 4, cf.cash_date, bg=bg, align="center")
        val(ws, row, 5, cf.cf_type,   bg=bg)
        val(ws, row, 6, cf.fund_amt,  "#,##0.000", bg=bg, align="right")
    borders(ws, cf_start-1, row, 2, 6)
    row += 2

    # ── NAV Bridge per fund ────────────────────────────────────────────────
    sec(ws, row, 2, "NAV Bridge (8-line, fund-level USD M)  ·  Reporting Period: FY 2024")
    row += 1

    bridge_labels = [
        ("beginning_nav",         "Beginning NAV",         False, LBGBLU),
        ("contribution",          "(+) Contribution",      False, None),
        ("disposition",           "(−) Disposition",       False, None),
        ("income",                "(+) Income",            False, None),
        ("cashflow_adjusted_nav", "= CF Adjusted NAV",     True,  LBGBLU),
        ("income_reversal",       "(−) Income Reversal",   False, None),
        ("write_up_down",         "(+/−) Write Up/Down *", False, YELLOW),
        ("ending_nav",            "= Ending NAV",          True,  LBGBLU),
    ]

    fund_ids = list(store.funds.keys())
    # header row
    hdr(ws, row, 2, "Line Item")
    for j, fid in enumerate(fund_ids):
        hdr(ws, row, 3+j, store.funds[fid].fund_name, wrap=True)
    ws.row_dimensions[row].height = 30
    bridge_hdr_row = row
    row += 1

    for key, label_text, bold, bg in bridge_labels:
        for j, fid in enumerate(fund_ids):
            bridge = answers.get(f"nav_bridge_{fid}", {})
            v = bridge.get(key, "")
            val(ws, row, 3+j, v, "#,##0.0000", bold=bold, bg=bg, align="right")
        val(ws, row, 2, label_text, bold=bold, bg=bg)
        row += 1

    borders(ws, bridge_hdr_row, row-1, 2, 2+len(fund_ids))
    ws.cell(row, 2, "* Write Up/Down = Ending NAV − (CF Adj NAV + Income Reversal)  ← plug, never directly input").font = \
        Font(italic=True, size=9, color="888888")
    row += 2

    # ── Performance Metrics ────────────────────────────────────────────────
    sec(ws, row, 2, "Performance Metrics  (visible by task level)")
    row += 1
    hdr(ws, row, 2, "Metric")
    hdr(ws, row, 3, "Visible at Level")
    for j, fid in enumerate(fund_ids):
        hdr(ws, row, 4+j, store.funds[fid].fund_name, wrap=True)
    ws.row_dimensions[row].height = 30
    met_hdr_row = row
    row += 1

    for key, lbl, level_label, fmt in [
        ("moic", "MOIC", "Level 2+", '0.000"x"'),
        ("irr",  "IRR",  "Level 3",  "0.0%"),
    ]:
        val(ws, row, 2, lbl, bold=True)
        val(ws, row, 3, level_label, bg=LBGGRN, align="center")
        for j, fid in enumerate(fund_ids):
            m = answers.get(f"metrics_{fid}", {})
            val(ws, row, 4+j, m.get(key, ""), fmt, align="right")
        row += 1
    borders(ws, met_hdr_row, row-1, 2, 3+len(fund_ids))
    row += 2

    # ── MOIC formula note ─────────────────────────────────────────────────
    note = ws.cell(row, 2, "MOIC = (Total Disposition + Total Income + Ending NAV) / Total Contributions")
    note.font = Font(italic=True, size=9, color="444444")
    row += 1
    note2 = ws.cell(row, 2, "IRR  = XIRR on all dated USD cashflows (contributions negative) + terminal Ending NAV at reporting date")
    note2.font = Font(italic=True, size=9, color="444444")


# ══════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

for task_id, loader in [
    ("easy",   load_easy_task),
    ("medium", load_medium_task),
    ("hard",   load_hard_task),
]:
    build_task_sheet(wb, task_id, loader)

path = "E:/Claude projects/Hackathon/FundLens Sample Data.xlsx"
wb.save(path)
print(f"Saved: {path}")
